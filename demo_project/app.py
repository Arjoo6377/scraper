from flask import Flask, render_template, send_file, jsonify
import requests
import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

def fetch_api_data():
    url = "https://polling.crisil.com/gateway/pollingsebi/api/amfi/fundperformance"
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Content-Type': 'application/json',
        'Origin': 'https://polling.crisil.com',
        'Referer': 'https://polling.crisil.com/polling/amfi/fund-performance',
        'User-Agent': 'Mozilla/5.0',
        'Authorization': '14210f2d-2c13-4bfd-92ac-beb41d0b5204'
    }
    payloads = [
        {"maturityType": i, "category": j, "subCategory": k, "mfid": 0, "reportDate": "23-Apr-2025"}
        for i in range(1, 5) for j in range(1, 5) for k in range(1, 5)
    ]
    all_data = []
    for payload in payloads:
        try:
            response = requests.post(url, headers=headers, json=payload)
            response.raise_for_status()
            data = response.json()
            if data and 'data' in data and data['data']:
                all_data.extend(data['data'])
        except:
            continue
    return all_data

def save_to_excel(data):
    try:
        column_name_map = {
            'schemeName': 'Scheme Name',
            'benchmark': 'Benchmark',
            'riskometerScheme': 'Riskometer Scheme',
            'riskometerBenchmark': 'Riskometer Benchmark',
            'navDate': 'NAV Date',
            'navRegular': 'NAV Regular',
            'navDirect': 'NAV Direct',
            'return1YearRegular': 'Return 1 Year (%) Regular',
            'return1YearDirect': 'Return 1 Year (%) Direct',
            'return1YearBenchmark': 'Return 1 Year (%) Benchmark',
            'ir1YrRegular': 'Information Ratio* 1 Year (Regular)',
            'ir1YrDirect': 'Information Ratio* 1 Year (Direct)',
            'return3YearRegular': 'Return 3 Year (%) Regular',
            'return3YearDirect': 'Return 3 Year (%) Direct',
            'return3YearBenchmark': 'Return 3 Year (%) Benchmark',
            'ir3YrRegular': 'Information Ratio* 3 Year (Regular)',
            'ir3YrDirect': 'Information Ratio* 3 Year (Direct)',
            'return5YearRegular': 'Return 5 Year (%) Regular',
            'return5YearDirect': 'Return 5 Year (%) Direct',
            'return5YearBenchmark': 'Return 5 Year (%) Benchmark',
            'ir5YrRegular': 'Information Ratio* 5 Year (Regular)',
            'ir5YrDirect': 'Information Ratio* 5 Year (Direct)',
            'return10YearRegular': 'Return 10 Year (%) Regular',
            'return10YearDirect': 'Return 10 Year (%) Direct',
            'return10YearBenchmark': 'Return 10 Year (%) Benchmark',
            'ir10YrRegular': 'Information Ratio* 10 Year (Regular)',
            'ir10YrDirect': 'Information Ratio* 10 Year (Direct)',
            'returnSinceLaunchRegular': 'Return Since Launch Regular',
            'returnSinceLaunchDirect': 'Return Since Launch Direct',
            'returnSinceLaunchBenchmarkRegular': 'Return Since Launch  Benchmark',
            'returnSinceLaunchBenchmarkDirect': 'Return Since Launch Direct Benchmark',
            'dailyAUM': 'Daily AUM (Cr.)'
        }

        columns_order = list(column_name_map.keys())

        df = pd.DataFrame(data)[columns_order]
        df = df.sort_values('schemeName')
        df = df.rename(columns=column_name_map)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"fund_performance_{timestamp}.xlsx"

        os.makedirs('static', exist_ok=True)
        excel_path = os.path.join('static', excel_filename)

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Fund Performance', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Fund Performance']

            header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            for col in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col)
                cell = worksheet[f"{column_letter}1"]
                cell.fill = header_fill
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                max_len = max(len(str(cell.value)), df[df.columns[col - 1]].astype(str).map(len).max())
                worksheet.column_dimensions[column_letter].width = min(max_len + 2, 30)

            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet[f"{get_column_letter(col)}{row}"]
                    cell.font = Font(name='Arial', size=10)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left' if col <= 4 else 'center', vertical='center')
                    if isinstance(cell.value, (int, float)):
                        col_name = df.columns[col - 1].lower()
                        if 'return' in col_name:
                            cell.number_format = '0.00%'
                        elif 'nav' in col_name or 'aum' in col_name:
                            cell.number_format = '#,##0.00'
                        elif 'ir' in col_name:
                            cell.number_format = '0.000'

            worksheet.freeze_panes = 'A2'

        return excel_filename
    except Exception as e:
        print(f"Excel Error: {str(e)}")
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate_excel')
def generate_excel():
    try:
        data = fetch_api_data()
        if not data:
            return jsonify({'success': False, 'error': 'No data received from API'})
        filename = save_to_excel(data)
        if filename:
            return jsonify({'success': True, 'filename': filename, 'record_count': len(data)})
        return jsonify({'success': False, 'error': 'Failed to generate Excel file'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/download/<filename>')
def download_file(filename):
    try:
        return send_file(os.path.join('static', filename), as_attachment=True, download_name=filename)
    except Exception as e:
        return str(e)

if __name__ == '__main__':
    app.run(debug=True)
